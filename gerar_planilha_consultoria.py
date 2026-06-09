import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Paleta de cores
COR_TITULO_GERAL    = "1F3864"  # azul escuro
COR_CABECALHO_ABA   = "2E75B6"  # azul médio
COR_CATEGORIA       = "BDD7EE"  # azul claro
COR_LINHA_PAR       = "F2F7FB"  # azul bem claro
COR_LINHA_IMPAR     = "FFFFFF"  # branco
COR_TEXTO_BRANCO    = "FFFFFF"
COR_TEXTO_ESCURO    = "1F3864"

borda_fina = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)

def estilo_titulo(ws, linha, texto, colunas_mesclar=None):
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = Font(bold=True, size=14, color=COR_TEXTO_BRANCO, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COR_TITULO_GERAL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if colunas_mesclar:
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha, end_column=colunas_mesclar)

def estilo_cabecalho(ws, linha, colunas):
    labels = ["#", "Categoria", "Pergunta", "Resposta / Observação", "Prioridade", "Status"]
    for col, label in enumerate(labels[:colunas], start=1):
        c = ws.cell(row=linha, column=col, value=label)
        c.font = Font(bold=True, size=11, color=COR_TEXTO_BRANCO, name="Calibri")
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO_ABA)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina

def estilo_categoria(ws, linha, texto, num_colunas):
    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha, end_column=num_colunas)
    c = ws.cell(row=linha, column=1, value=f"  {texto}")
    c.font = Font(bold=True, size=11, color=COR_TEXTO_ESCURO, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COR_CATEGORIA)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = borda_fina

def linha_pergunta(ws, linha, num, categoria, pergunta, prioridade="Alta"):
    par = (linha % 2 == 0)
    cor_bg = COR_LINHA_PAR if par else COR_LINHA_IMPAR
    dados = [num, categoria, pergunta, "", prioridade, "Pendente"]
    for col, val in enumerate(dados, start=1):
        c = ws.cell(row=linha, column=col, value=val)
        c.font = Font(size=10, name="Calibri", color="1A1A1A")
        c.fill = PatternFill("solid", fgColor=cor_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = borda_fina
        if col == 3:  # coluna pergunta
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if col == 4:  # coluna resposta
            c.fill = PatternFill("solid", fgColor="FFFDE7")
        if col == 5:  # prioridade
            cores_prioridade = {"Alta": "FF6B6B", "Média": "FFB347", "Baixa": "98D8A3"}
            cor_p = cores_prioridade.get(prioridade, "CCCCCC")
            c.fill = PatternFill("solid", fgColor=cor_p)
            c.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
        if col == 6:  # status
            c.fill = PatternFill("solid", fgColor="E8E8E8")
            c.alignment = Alignment(horizontal="center", vertical="center")

# ============================================================
# DADOS DAS ABAS
# ============================================================

abas = {
    "1. Diagnóstico Geral": {
        "cor": "1F3864",
        "categorias": [
            ("DADOS CADASTRAIS DA EMPRESA", "Alta", [
                "Qual a razão social completa e CNPJ da Matriz (Senador Lemos)?",
                "Qual a razão social completa e CNPJ da Filial (BR 316)?",
                "Quais os endereços completos de cada unidade?",
                "Qual o horário de funcionamento da Matriz e da Filial?",
                "Há quanto tempo a empresa está no mercado?",
                "Qual o regime tributário atual (Simples Nacional, Lucro Presumido, Lucro Real)?",
                "A empresa possui alvarás e licenças em dia?",
                "Tem seguro empresarial?",
            ]),
            ("PERFIL DO EMPRESÁRIO", "Alta", [
                "Qual a formação e experiência do Alexandre no setor?",
                "Quantas horas por dia trabalha em média?",
                "Como divide o tempo entre Matriz e Filial?",
                "Tem sócios ativos na operação diária? Qual o papel de cada um?",
                "Quais decisões somente o Alexandre pode tomar?",
                "O que mais consome seu tempo no dia a dia?",
                "Quais tarefas gostaria de eliminar da sua rotina?",
                "Consegue tirar férias sem que a operação sofra?",
                "Tem vida pessoal preservada fora do trabalho?",
                "Qual é a sua maior dor/frustração com o negócio hoje?",
            ]),
            ("VISÃO E OBJETIVOS", "Alta", [
                "Qual é a missão e visão da Fácil Construir?",
                "Quais são as metas para os próximos 12 meses?",
                "Quais são os maiores desafios atuais?",
                "Planeja abrir novas unidades?",
                "Tem plano de saída ou sucessão para o futuro?",
                "O que mudaria imediatamente se tivesse os recursos?",
                "Qual a visão de 5 anos para o negócio?",
                "Quer trabalhar menos e ganhar mais? Qual o equilíbrio ideal?",
            ]),
        ]
    },
    "2. Financeiro": {
        "cor": "2E75B6",
        "categorias": [
            ("CONTROLE FINANCEIRO", "Alta", [
                "Utiliza sistema de gestão (ERP)? Qual?",
                "Controla fluxo de caixa diariamente?",
                "Tem contador ou escritório contábil? Qual a qualidade do serviço?",
                "Recebe DRE (Demonstrativo de Resultado) mensalmente?",
                "Sabe qual é a margem de lucro líquida da empresa?",
                "Controla custos fixos e variáveis separadamente?",
                "Tem caixa separado por unidade (Matriz e Filial)?",
                "As finanças pessoais e da empresa são separadas?",
                "Tem pró-labore definido?",
                "Qual o faturamento médio mensal da Matriz?",
                "Qual o faturamento médio mensal da Filial?",
                "Qual o ticket médio por venda em cada unidade?",
            ]),
            ("RECEBIMENTOS E PAGAMENTOS", "Alta", [
                "Quais as formas de pagamento aceitas (dinheiro, cartão, PIX, boleto)?",
                "Tem controle de contas a receber?",
                "Qual o prazo médio de recebimento?",
                "Qual a inadimplência média atual?",
                "Tem política de crédito para clientes definida?",
                "Quem aprova crédito para clientes?",
                "Qual o prazo médio de pagamento a fornecedores?",
                "Tem capital de giro suficiente para a operação?",
                "Tem acesso a linhas de crédito/financiamento para estoque?",
                "Controla contas a pagar com antecedência?",
            ]),
            ("INDICADORES FINANCEIROS", "Média", [
                "Sabe qual é o ponto de equilíbrio (break-even) de cada loja?",
                "Acompanha o retorno sobre investimento (ROI)?",
                "Tem metas de faturamento mensais definidas?",
                "Compara resultados mês a mês?",
                "Tem reserva de emergência para o negócio?",
                "Faz planejamento financeiro anual?",
                "Conhece quais produtos têm maior e menor margem?",
            ]),
        ]
    },
    "3. Estoque e Compras": {
        "cor": "2E75B6",
        "categorias": [
            ("CONTROLE DE ESTOQUE", "Alta", [
                "Como é feito o controle de estoque atualmente (sistema, planilha, manual)?",
                "Com que frequência realiza inventário?",
                "O estoque é integrado entre Matriz e Filial?",
                "Tem produtos com giro lento ou encalhados?",
                "Tem problemas de ruptura de estoque (falta de produto)?",
                "Quem é responsável pelo controle de estoque em cada unidade?",
                "Tem espaço físico suficiente para estoque nas duas unidades?",
                "Controla devoluções e avarias?",
                "Usa código de barras ou etiquetagem dos produtos?",
            ]),
            ("GESTÃO DE COMPRAS", "Alta", [
                "Quem faz as compras? O Alexandre ou tem comprador dedicado?",
                "Como decide o que comprar e quando comprar?",
                "Tem ponto de pedido definido por produto?",
                "Quais são as principais linhas/categorias de produtos?",
                "Quais os 10 principais fornecedores?",
                "Com que frequência os representantes visitam a loja?",
                "Tem calendário de negociação com fornecedores?",
                "Negocia bonificações, verbas de marketing ou descontos progressivos?",
                "Tem exclusividade com algum fornecedor ou marca?",
                "Qual o prazo médio de entrega dos fornecedores?",
                "Tem problemas de qualidade com algum fornecedor?",
                "Participa de feiras do setor (Feicon, etc.)?",
            ]),
            ("LOGÍSTICA E ENTREGA", "Média", [
                "Oferece serviço de entrega para clientes?",
                "Tem frota própria ou terceiriza?",
                "Qual o raio de entrega?",
                "Cobra pela entrega?",
                "Controla prazo de entrega ao cliente?",
                "Tem política de troca e devolução definida?",
            ]),
        ]
    },
    "4. Vendas e Atendimento": {
        "cor": "2E75B6",
        "categorias": [
            ("EQUIPE DE VENDAS", "Alta", [
                "Quantos vendedores tem em cada unidade?",
                "Os vendedores têm metas definidas?",
                "Como são remunerados (fixo + comissão)? Qual o percentual?",
                "Os vendedores têm treinamento de produto e técnica de vendas?",
                "Tem alta rotatividade na equipe de vendas?",
                "Quem é o melhor vendedor e por quê?",
                "Os vendedores fazem atendimento externo (obra, projeto)?",
                "Tem script ou processo de vendas padronizado?",
            ]),
            ("PROCESSO DE VENDA", "Alta", [
                "Como é o fluxo do cliente na loja (da entrada até a saída)?",
                "Faz orçamentos? Como controla e acompanha?",
                "Qual a taxa de conversão de orçamentos em vendas?",
                "Qual o prazo médio de entrega de orçamento ao cliente?",
                "Tem processo de pós-venda?",
                "Como trata reclamações de clientes?",
                "Qual a proporção de clientes PF (pessoa física) x PJ (construtoras, pedreiros)?",
                "Tem canal de vendas online (WhatsApp, site, e-commerce)?",
                "Faz venda por telefone ou WhatsApp sem o cliente vir à loja?",
            ]),
            ("CLIENTES E FIDELIZAÇÃO", "Alta", [
                "Tem base de dados de clientes cadastrados?",
                "Quantos clientes ativos tem na base?",
                "Sabe quais clientes não voltaram nos últimos 90 dias?",
                "Tem programa de fidelidade ou benefícios para clientes recorrentes?",
                "Faz distinção no atendimento para clientes VIP?",
                "Conhece os 20 maiores clientes por faturamento?",
                "Clientes B2B (construtoras, pedreiros) têm condições especiais?",
                "Como capta novos clientes hoje?",
            ]),
        ]
    },
    "5. Marketing": {
        "cor": "2E75B6",
        "categorias": [
            ("PRESENÇA DIGITAL", "Alta", [
                "Tem perfil no Google Meu Negócio (Maps) atualizado?",
                "Qual a avaliação atual no Google (estrelas) das duas unidades?",
                "Tem site? Está atualizado?",
                "Tem perfil no Instagram? Quantidade de seguidores?",
                "Tem perfil no Facebook? Quantidade de seguidores?",
                "Tem canal no YouTube ou TikTok?",
                "Com que frequência posta nas redes sociais?",
                "Quem cuida das redes sociais hoje?",
                "Tem identidade visual definida (logo, cores, tipografia)?",
                "Tem cadastro no iFood Mercado, OLX ou similares?",
            ]),
            ("INVESTIMENTO EM MARKETING", "Alta", [
                "Qual o investimento mensal em marketing atualmente?",
                "Faz anúncios pagos (Google Ads, Meta Ads / Facebook e Instagram)?",
                "Já tentou tráfego pago antes? Qual foi o resultado?",
                "Faz panfletagem, outdoor ou marketing offline?",
                "Tem parcerias com construtoras, arquitetos ou pedreiros?",
                "Faz campanhas sazonais (Dia das Mães, Black Friday, etc.)?",
                "Tem materiais gráficos padronizados (banner, cartão de visita)?",
            ]),
            ("COMUNICAÇÃO COM CLIENTES", "Alta", [
                "Faz disparo de mensagens para base de clientes (WhatsApp, SMS, e-mail)?",
                "Com que frequência comunica promoções para a base?",
                "Tem diferencial competitivo comunicado claramente ao mercado?",
                "Como está posicionado frente à concorrência?",
                "Quais são os principais concorrentes na região?",
                "Tem proposta de valor única definida?",
                "Clientes indicam outros clientes (programa de indicação)?",
            ]),
        ]
    },
    "6. Tecnologia e Ferramentas": {
        "cor": "2E75B6",
        "categorias": [
            ("SISTEMAS DE GESTÃO", "Alta", [
                "Qual sistema ERP/PDV utiliza (Bling, Tiny, TOTVS, Linx, outro)?",
                "O sistema atende às necessidades? Quais as principais limitações?",
                "O sistema integra Matriz e Filial?",
                "Emite NF-e e NF-c regularmente pelo sistema?",
                "Tem backup dos dados?",
                "Usa sistema de ponto eletrônico para funcionários?",
                "Tem sistema de câmeras com acesso remoto?",
            ]),
            ("COMUNICAÇÃO E FERRAMENTAS DIGITAIS", "Alta", [
                "Usa WhatsApp Business (não pessoal) nas lojas?",
                "Tem WhatsApp API (número fixo com multi-atendimento)?",
                "Usa algum CRM? Qual?",
                "Usa Google Workspace (Gmail, Drive, Sheets) ou Microsoft 365?",
                "A internet das lojas é estável e rápida?",
                "Tem computadores/tablets disponíveis para vendedores?",
                "Usa algum aplicativo de gestão no celular?",
                "Usa ferramenta de comunicação interna com a equipe (grupos, app)?",
                "Tem e-mail profissional (não @gmail pessoal)?",
            ]),
            ("AUTOMAÇÃO E IA", "Alta", [
                "Já usa alguma ferramenta de automação (Zapier, Make, n8n)?",
                "Já teve experiência com chatbot ou atendimento automatizado?",
                "Qual a abertura do Alexandre para adotar novas tecnologias?",
                "A equipe tem resistência ao uso de tecnologia?",
                "Quais tarefas repetitivas gostaria de automatizar primeiro?",
                "Prefere receber relatórios por WhatsApp, e-mail ou dashboard?",
                "Tem interesse em usar IA para criar conteúdo (posts, respostas)?",
                "Quer automação de pesquisa de concorrentes e fornecedores?",
            ]),
        ]
    },
    "7. CRM e Follow-up": {
        "cor": "2E75B6",
        "categorias": [
            ("BASE DE CLIENTES E CRM", "Alta", [
                "Tem base de clientes com nome, telefone e e-mail cadastrados?",
                "A base está organizada e segmentada?",
                "Quantos contatos tem na base (estimativa)?",
                "Usa alguma ferramenta de CRM (HubSpot, RD Station, PipeDrive, outro)?",
                "O histórico de compras por cliente está disponível?",
                "Tem segmentação por tipo de cliente (PF, pedreiro, construtora, arquiteto)?",
                "Tem segmentação por produto comprado ou obra?",
            ]),
            ("FOLLOW-UP E REATIVAÇÃO", "Alta", [
                "Faz follow-up após envio de orçamento?",
                "Qual o prazo para follow-up de orçamentos?",
                "Quem faz o follow-up hoje?",
                "Faz ação para reativar clientes inativos?",
                "Sabe identificar clientes que estão comprando menos?",
                "Tem régua de relacionamento (sequência de mensagens) definida?",
                "Faz pesquisa de satisfação (NPS) com clientes?",
                "Como trata clientes que reclamaram? Tem processo?",
            ]),
            ("ATENDIMENTO NO WHATSAPP", "Alta", [
                "Quantas mensagens por dia recebe no WhatsApp da loja?",
                "Quem responde o WhatsApp atualmente?",
                "Qual o tempo médio de resposta?",
                "Perde vendas por demora no atendimento pelo WhatsApp?",
                "Tem catálogo de produtos configurado no WhatsApp Business?",
                "Usa respostas rápidas no WhatsApp Business?",
                "Tem mensagem automática de ausência configurada?",
                "Quer implementar chatbot para atendimento inicial?",
                "Aceita que a IA responda dúvidas básicas automaticamente?",
                "Quais as perguntas mais frequentes dos clientes no WhatsApp?",
            ]),
        ]
    },
    "8. Rotinas Administrativas": {
        "cor": "2E75B6",
        "categorias": [
            ("ROTINAS DIÁRIAS", "Alta", [
                "Descreva a rotina diária de abertura da loja.",
                "Descreva a rotina diária de fechamento da loja.",
                "Quem faz o fechamento de caixa diário?",
                "Tem checklist de abertura e fechamento?",
                "Como é feita a conferência de estoque diária?",
                "Quem cuida da organização e limpeza das lojas?",
                "Como é tratada a reposição de produtos nas prateleiras?",
                "Tem reunião diária (mesmo que rápida) com a equipe?",
            ]),
            ("ROTINAS SEMANAIS E MENSAIS", "Alta", [
                "Tem reunião semanal de vendas? Quem participa?",
                "Faz análise semanal de indicadores (vendas, estoque)?",
                "Quando faz pedidos de compra? Qual a frequência?",
                "Faz conferência de contas a pagar/receber semanalmente?",
                "Tem reunião mensal de resultados?",
                "Quando envia relatórios ao contador?",
                "Tem processo de avaliação mensal de funcionários?",
                "Faz inventário parcial ou total mensalmente?",
            ]),
            ("GESTÃO DE PESSOAS E RH", "Média", [
                "Tem organograma definido para cada unidade?",
                "Quem é o gerente/responsável de cada loja?",
                "Como é o processo de admissão de novos funcionários?",
                "Faz treinamentos de integração para novos colaboradores?",
                "Tem alto índice de rotatividade? Por quê?",
                "Como comunica novidades e mudanças para a equipe?",
                "Tem plano de cargos e salários?",
                "Os funcionários conhecem as metas da empresa?",
                "Tem indicadores de desempenho por colaborador?",
            ]),
        ]
    },
    "9. Assistente Virtual": {
        "cor": "2E75B6",
        "categorias": [
            ("PROSPECÇÃO ATIVA", "Alta", [
                "Quais segmentos quer prospectar ativamente (construtoras, prefeitura, escolas)?",
                "Quais bairros ou cidades quer prospectar?",
                "Quer prospectar pelo Google Maps (pesquisa de CNPJ)?",
                "Tem lista de obras em andamento na região para prospectar?",
                "Quer prospectar arquitetos e designers de interiores?",
                "Quais produtos ou linhas quer priorizar na prospecção?",
                "Prefere prospecção por WhatsApp, e-mail ou ligação?",
                "Tem equipe para fazer follow-up das prospecções?",
            ]),
            ("ASSISTENTE PARA O ALEXANDRE", "Alta", [
                "Quais relatórios quer receber automaticamente (diário, semanal)?",
                "Quer receber resumo diário de vendas pelo WhatsApp?",
                "Quer alertas de estoque baixo?",
                "Quer resumo de aniversariantes do dia para mensagem de parabéns?",
                "Quer agenda de compromissos e reuniões organizada?",
                "Quer controle de metas diárias com notificação?",
                "Quer assistente para pesquisar preços de concorrentes?",
                "Quer auxílio para criação de conteúdo para redes sociais?",
                "Quer IA para responder e-mails ou mensagens padrão?",
                "Quais decisões quer ter dados na hora para decidir mais rápido?",
            ]),
            ("ATENDIMENTO AUTOMATIZADO", "Alta", [
                "Quer chatbot para triagem inicial no WhatsApp?",
                "Quer automação para envio de orçamentos?",
                "Quer resposta automática para perguntas frequentes?",
                "Quer fluxo automático de follow-up de orçamentos?",
                "Quer automação de pesquisa de satisfação pós-venda?",
                "Quer sistema de agendamento de visitas técnicas automatizado?",
                "Quer integração WhatsApp com sistema ERP/CRM?",
                "Quer disparo automático de promoções segmentadas?",
            ]),
        ]
    },
    "10. Plano de Marketing": {
        "cor": "2E75B6",
        "categorias": [
            ("DIAGNÓSTICO DE MARKETING", "Alta", [
                "Qual é o público-alvo principal de cada unidade?",
                "Qual a percepção de marca da Fácil Construir na região?",
                "O que os clientes mais elogiam?",
                "O que os clientes mais reclamam?",
                "Quais produtos são destaque e geram mais interesse?",
                "Quais produtos têm maior margem e deveriam ser mais promovidos?",
                "Como a concorrência se comunica e se posiciona?",
                "Tem diferencial competitivo claro e comunicado?",
            ]),
            ("ESTRATÉGIA DE CONTEÚDO", "Alta", [
                "Tem câmera ou smartphone com boa câmera para fotos/vídeos?",
                "Tem funcionário com perfil para aparecer nas redes?",
                "Quer Alexandre aparecer como porta-voz da marca?",
                "Quais temas de conteúdo fazem sentido (dicas de obra, produtos, promoções)?",
                "Com que frequência quer publicar (Instagram, Facebook, YouTube)?",
                "Quer fazer reels/shorts para alcance orgânico?",
                "Quer fazer lives de ofertas ou demonstração de produtos?",
                "Tem interesse em parcerias com influenciadores locais?",
            ]),
            ("TRÁFEGO PAGO E CAMPANHAS", "Alta", [
                "Qual o orçamento disponível para tráfego pago mensalmente?",
                "Quer anunciar no Google (pesquisa e display)?",
                "Quer anunciar no Instagram/Facebook?",
                "Tem pixels configurados no site?",
                "Quer remarketing para visitantes do site?",
                "Quer campanhas específicas para cada unidade?",
                "Tem datas sazonais prioritárias para campanhas (Dia do Pedreiro, etc.)?",
                "Quer mensurar retorno de cada campanha (ROI)?",
            ]),
            ("GOOGLE MEU NEGÓCIO E LOCAL SEO", "Alta", [
                "Tem acesso às contas do Google Meu Negócio das duas unidades?",
                "As informações (horário, endereço, telefone) estão corretas?",
                "Tem fotos profissionais das lojas no Google?",
                "Responde avaliações no Google?",
                "Tem estratégia para solicitar avaliações de clientes satisfeitos?",
                "Aparece nas pesquisas locais ('material de construção perto de mim')?",
                "Quer otimizar o SEO local das duas unidades?",
            ]),
        ]
    },
}


# ============================================================
# GERAR ABAS
# ============================================================

for idx, (nome_aba, dados) in enumerate(abas.items()):
    if idx == 0:
        ws = wb.active
        ws.title = nome_aba
    else:
        ws = wb.create_sheet(title=nome_aba)

    ws.sheet_view.showGridLines = False

    # Larguras das colunas
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    # Título
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 28
    estilo_titulo(ws, 1,
        f"CONSULTORIA FÁCIL CONSTRUIR  |  Empresário: Alexandre  |  {nome_aba.replace('1. ','').replace('2. ','').replace('3. ','').replace('4. ','').replace('5. ','').replace('6. ','').replace('7. ','').replace('8. ','').replace('9. ','').replace('10. ','')}",
        6)
    ws.row_dimensions[2].height = 6
    estilo_cabecalho(ws, 3, 6)

    linha_atual = 4
    contador = 1

    for categoria, prioridade_padrao, perguntas in dados["categorias"]:
        ws.row_dimensions[linha_atual].height = 22
        estilo_categoria(ws, linha_atual, f"▶  {categoria}", 6)
        linha_atual += 1

        for i, pergunta in enumerate(perguntas):
            ws.row_dimensions[linha_atual].height = 28
            linha_pergunta(ws, linha_atual, contador,
                           categoria[:25] + "...", pergunta, prioridade_padrao)
            linha_atual += 1
            contador += 1

    # Congelar painel no cabeçalho
    ws.freeze_panes = ws["A4"]

    # Filtro automático
    ws.auto_filter.ref = f"A3:F{linha_atual - 1}"


# ============================================================
# ABA CAPA / RESUMO
# ============================================================

capa = wb.create_sheet(title="CAPA", index=0)
capa.sheet_view.showGridLines = False
capa.column_dimensions["A"].width = 2
capa.column_dimensions["B"].width = 60
capa.column_dimensions["C"].width = 30

def cell_capa(row, col, val, bold=False, size=11, cor_bg=None, cor_fonte="1A1A1A", halign="left"):
    c = capa.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=size, name="Calibri", color=cor_fonte)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    if cor_bg:
        c.fill = PatternFill("solid", fgColor=cor_bg)
    return c

capa.row_dimensions[1].height = 10
capa.row_dimensions[2].height = 55
capa.merge_cells("B2:C2")
c = capa.cell(row=2, column=2, value="CONSULTORIA FÁCIL CONSTRUIR")
c.font = Font(bold=True, size=22, name="Calibri", color=COR_TEXTO_BRANCO)
c.fill = PatternFill("solid", fgColor=COR_TITULO_GERAL)
c.alignment = Alignment(horizontal="center", vertical="center")

capa.row_dimensions[3].height = 30
capa.merge_cells("B3:C3")
c = capa.cell(row=3, column=2, value="Razão Social: Construir Comércio  |  Empresário: Alexandre")
c.font = Font(bold=False, size=13, name="Calibri", color=COR_TEXTO_BRANCO)
c.fill = PatternFill("solid", fgColor=COR_CABECALHO_ABA)
c.alignment = Alignment(horizontal="center", vertical="center")

capa.row_dimensions[4].height = 10

infos = [
    (5, "Matriz:", "Senador Lemos"),
    (6, "Filial:", "BR 316"),
    (7, "Objetivo:", "Otimizar rotinas, implantar assistente virtual, CRM, marketing e automações"),
    (8, "Consultor(a):", ""),
    (9, "Data de início:", ""),
    (10, "Versão:", "1.0"),
]

for row, label, valor in infos:
    capa.row_dimensions[row].height = 26
    cell_capa(row, 2, label, bold=True, size=11, cor_bg="EBF3FB", cor_fonte=COR_TEXTO_ESCURO)
    cell_capa(row, 3, valor, bold=False, size=11, cor_bg="FFFFFF")
    capa.cell(row=row, column=2).border = borda_fina
    capa.cell(row=row, column=3).border = borda_fina

capa.row_dimensions[11].height = 16

capa.row_dimensions[12].height = 26
capa.merge_cells("B12:C12")
cell_capa(12, 2, "ABAS DO DIAGNÓSTICO", bold=True, size=12,
          cor_bg=COR_CABECALHO_ABA, cor_fonte=COR_TEXTO_BRANCO, halign="center")

lista_abas = [
    ("1. Diagnóstico Geral",   "Dados da empresa, perfil do empresário, visão e objetivos"),
    ("2. Financeiro",          "Controle financeiro, recebimentos, pagamentos e indicadores"),
    ("3. Estoque e Compras",   "Gestão de estoque, compras, fornecedores e logística"),
    ("4. Vendas e Atendimento","Equipe, processo de venda, clientes e fidelização"),
    ("5. Marketing",           "Presença digital, investimento em marketing, comunicação"),
    ("6. Tecnologia",          "Sistemas ERP, WhatsApp Business, automação e IA"),
    ("7. CRM e Follow-up",     "Base de clientes, follow-up, reativação e WhatsApp"),
    ("8. Rotinas Adm.",        "Rotinas diárias, semanais, mensais e gestão de pessoas"),
    ("9. Assistente Virtual",  "Prospecção ativa, assistente pessoal e atendimento automatizado"),
    ("10. Plano de Marketing", "Diagnóstico, conteúdo, tráfego pago e Google Meu Negócio"),
]

for i, (nome, descricao) in enumerate(lista_abas):
    row = 13 + i
    capa.row_dimensions[row].height = 24
    c1 = capa.cell(row=row, column=2, value=nome)
    c1.font = Font(bold=True, size=10, name="Calibri", color=COR_TEXTO_ESCURO)
    c1.fill = PatternFill("solid", fgColor=COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = borda_fina

    c2 = capa.cell(row=row, column=3, value=descricao)
    c2.font = Font(size=10, name="Calibri", color="444444")
    c2.fill = PatternFill("solid", fgColor=COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = borda_fina


# ============================================================
# SALVAR
# ============================================================

caminho = "/home/user/curso-financeiro/Consultoria_FacilConstruir_Diagnostico.xlsx"
wb.save(caminho)
print(f"Planilha gerada: {caminho}")
